from kor.extraction import create_extraction_chain
from kor.nodes import Object, Text, Number

# LangChain Models
from langchain_core.prompts import PromptTemplate
from langchain_openai import OpenAI
from langchain_openai import ChatOpenAI

# Standard Helpers
import time
import json
from datetime import datetime

# For token counting
from langchain.callbacks import get_openai_callback

import os
import openpyxl

llm = ChatOpenAI(
    model_name=gpt,
    temperature=0, # Temperature sets how random the choice of tokens is. This means that the lower the temperature, the least creativity the model has
    max_tokens=2000,
    openai_api_key=api_key
)


chain = create_extraction_chain(llm, instruction, encoder_or_encoder_class="csv")
def load_processed_files(log_file):
    """Load the list of processed files from the log file."""
    if os.path.exists(log_file):
        with open(log_file, 'r') as f:
            return set(line.strip() for line in f)
    return set()

def update_processed_files(log_file, processed_files):
    """Update the log file with the list of processed files."""
    with open(log_file, 'a') as f:
        for filename in processed_files:
            f.write(f"{filename}\n")

def handle_value(value):
    """Convert value to a string and handle lists or empty values."""
    if isinstance(value, list):
        return ', '.join(map(str, value))  # Convert each list item to string and join
    return value if value else "NA"

def process_files_in_folder(summaries_out, results_file, errors_file, log_file):
    # Load already processed files
    processed_files = load_processed_files(log_file)

    # Check if the workbook already exists
    if os.path.exists(results_file):
        # Load the existing workbook and sheets
        wb = openpyxl.load_workbook(results_file)
        results_sheet = wb.get_sheet_by_name('Results')
        errors_sheet = wb.get_sheet_by_name('Errors')
    else:
        # Create a new workbook and sheets for results and errors
        wb = openpyxl.Workbook()
        results_sheet = wb.active
        results_sheet.title = "Results"
        errors_sheet = wb.create_sheet(title="Errors")
        # Write headers to results sheet
        results_sheet.append(headers_vars)
        # Write header for errors sheet
        errors_sheet.append(["Error Files"])

    # List to keep track of error files
    error_files = []
    new_processed_files = set()  # Track files processed in this run

    # Get list of files in the folder and sort them alphabetically
    files = [f for f in os.listdir(summaries_out) if f.endswith(".txt")]
    files.sort()  # Sort files alphabetically


def handle_value(value):
    """Convert value to a string and handle lists or empty values."""
    if isinstance(value, list):
        return ', '.join(map(str, value))  # Convert each list item to string and join
    return value if value else "NA"

def pasqui_structuring(summaries_out, results_file, errors_file, log_file):
    # Load already processed files
    processed_files = load_processed_files(log_file)

    # Check if the workbook already exists
    if os.path.exists(results_file):
        # Load the existing workbook and sheets
        wb = openpyxl.load_workbook(results_file)
        results_sheet = wb.get_sheet_by_name('Results')
        errors_sheet = wb.get_sheet_by_name('Errors')
    else:
        # Create a new workbook and sheets for results and errors
        wb = openpyxl.Workbook()
        results_sheet = wb.active
        results_sheet.title = "Results"
        errors_sheet = wb.create_sheet(title="Errors")
        # Write headers to results sheet
        results_sheet.append(headers)
        # Write header for errors sheet
        errors_sheet.append(["Error Files"])

    # List to keep track of error files
    error_files = []
    new_processed_files = set()  # Track files processed in this run

    # Get list of files in the folder and sort them alphabetically
    files = [f for f in os.listdir(folder_path) if f.endswith(".txt")]
    files.sort()  # Sort files alphabetically

    # Process each file in the sorted list
    for filename in files:
        if filename not in processed_files:  # Ensure not to reprocess already processed files
            file_path = os.path.join(folder_path, filename)
            with open(file_path, 'r', encoding='utf-8') as file:
                text = file.read()
                input_data = {"text": text}
                try:
                    # Invoke the chain with the input data
                    output = chain.invoke(input_data)
                    print(f"Output for {filename}: {output}")  # Debugging line
                    data = output.get('data', {})
                    instruction_list = data.get("instruction", [])  # Get the list
                    print(f"instruction list for {filename}: {instruction_list}")  # Debugging line

                    if instruction_list:
                     for instruction in instruction_list:
                        row = [filename]

                        # Loop over the headers and dynamically fetch and process each field
                        for header in headers_vars[1:]:  # Skip the first element 'case_name'
                            value = instruction.get(header)
                            row.append(handle_value(value))

                        # Append the row with processed values to results_sheet
                        results_sheet.append(row)
                    else:
                    # Handle case where there are no crime records
                      row = [filename] + ["NA"] * (len(headers_vars) - 1)
                      results_sheet.append(row)

                    # Add file to the new processed files set
                    new_processed_files.add(filename)

                    # Save the workbook after processing each file
                    wb.save(results_file)

                    # Update the log file with the processed file
                    update_processed_files(log_file, {filename})

                except Exception as e:
                    # Handle any exceptions and store the error message
                    print(f"Error processing file {filename}: {e}")
                    error_files.append([filename])  # Record filename with error

    # Write error files to the errors sheet
    for filename in error_files:
        errors_sheet.append(filename)

    # Save the workbook after processing all files
    wb.save(results_file)

    # Update the log file with new processed files
    update_processed_files(log_file, new_processed_files)
    print(f"Results saved to {output_file_results}")
    print(f"Errors saved to {output_file_errors}")



