from.files import read_json_file,check_file_existence,check_create_folder,check_image_resolution,resize_image_resolution,read_setting,find_ori_files_from_json,delete_file,save_df_with_links,text_to_docx,split_path
from.image import resize_image_to_target_size,get_image_file_size
from.data_cleaning import linear_interpolation,check_for_discontinuities,merge_columns_to_dict_series,check_column_names
from.beiker_string import split_list_by_length,extract_str_by_pattern,extract_archive_class,remove_all_whitespaces,contains_empty_string,extract_str_from_pattern_in_json_list,is_labels_series,convert_path_to_system_style
import pandas as pd
from typing import List
from alibabacloud_ocr_api20210707.client import Client as ocr_api20210707Client
from alibabacloud_tea_openapi import models as open_api_models
from alibabacloud_ocr_api20210707 import models as ocr_api_20210707_models
from alibabacloud_tea_util import models as util_models
from alibabacloud_tea_console.client import Client as ConsoleClient
from alibabacloud_tea_util.client import Client as UtilClient
import os,re,io,json,datetime
from PIL import Image
from.log import printLog
from functools import partial  
from.safe import decrypt_simple
from openpyxl import load_workbook  
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
import paddle
from paddleocr import PaddleOCR
# from paddleocr import draw_ocr

# OCR开发文档：https://help.aliyun.com/zh/ocr/developer-reference/api-ocr-api-2021-07-07-recognizetableocr?spm=a2c4g.11186623.0.0.537b78eawsmonL
class Aliyun_table_Ocr:
    def __init__(self):
        pass

    @staticmethod
    def create_client(fp_setting) -> ocr_api20210707Client:
        """
        使用AK&SK初始化账号Client
        @return: Client
        @throws Exception
        """

        # 工程代码泄露可能会导致 AccessKey 泄露，并威胁账号下所有资源的安全性。以下代码示例仅供参考。
        # 建议使用更安全的 STS 方式，更多鉴权访问方式请参见：https://help.aliyun.com/document_detail/378659.html。
        config = open_api_models.Config(
            # 必填，请确保代码运行环境设置了环境变量 ALIBABA_CLOUD_ACCESS_KEY_ID。

            # 必填，请确保代码运行环境设置了环境变量 ALIBABA_CLOUD_ACCESS_KEY_SECRET。,

            access_key_id=decrypt_simple(read_setting(fp_setting,'OCR.ALIBABA_CLOUD_ACCESS_KEY_ID')),

            access_key_secret=decrypt_simple(read_setting(fp_setting,'OCR.ALIBABA_CLOUD_ACCESS_KEY_SECRET'))
        )
        # Endpoint 请参考 https://api.aliyun.com/product/ocr-api
        config.endpoint = f'ocr-api.cn-hangzhou.aliyuncs.com'
        return ocr_api20210707Client(config)

    @staticmethod
    def main(img,fp_setting) -> str: # args: List[str]  ->None
        client = Aliyun_table_Ocr.create_client(fp_setting)
        recognize_table_ocr_request = ocr_api_20210707_models.RecognizeTableOcrRequest(
            # image_url='http://viapi-test.oss-cn-shanghai.aliyuncs.com/viapi-3.0domepic/ocr/RecognizeTable/RecognizeTable1.jpg',

            body=img,
            is_hand_writing ="false",
            line_less= False,
            need_rotate = False,
            skip_detection=False
        )

        runtime = util_models.RuntimeOptions()
        try:
            resp = client.recognize_table_ocr_with_options(recognize_table_ocr_request, runtime)
            return UtilClient.to_jsonstring(resp)

            # ConsoleClient.log(type(resp)) # recognizeTableOcrResponse类型
            # ConsoleClient.log(type(UtilClient.to_jsonstring(resp))) #str类型

            # ConsoleClient.log(UtilClient.to_jsonstring(resp))
        except Exception as error:
            # 此处仅做打印展示，请谨慎对待异常处理，在工程项目中切勿直接忽略异常。
            # 错误 message
            printLog(error.message)
            # 诊断地址
            printLog(error.data.get("Recommend"))
            UtilClient.assert_as_string(error.message)

    @staticmethod
    async def main_async(args: List[str],fp_setting) -> None:
        client = Aliyun_table_Ocr.create_client(fp_setting)
        recognize_table_ocr_request = ocr_api_20210707_models.RecognizeTableOcrRequest(
            # image_url='http://viapi-test.oss-cn-shanghai.aliyuncs.com/viapi-3.0domepic/ocr/RecognizeTable/RecognizeTable1.jpg',

            body=img,
            is_hand_writing ="false",
            line_less= False,
            need_rotate = False,
            skip_detection=False
        )
        runtime = util_models.RuntimeOptions()
        try:
            resp = await client.recognize_table_ocr_with_options_async(recognize_table_ocr_request, runtime)
            ConsoleClient.log(UtilClient.to_jsonstring(resp))
        except Exception as error:
            # 此处仅做打印展示，请谨慎对待异常处理，在工程项目中切勿直接忽略异常。
            # 错误 message
            printLog(error.message)
            # 诊断地址
            printLog(error.data.get("Recommend"))
            UtilClient.assert_as_string(error.message)

def ali_tableocr_calculate_dimensions(cellInfos):
    """
    获取ali云ocr表格识别后的json文件中的行数和列数。ali云表格识别后json数据中的cellInfos如下：
    [{\"pos\":[{\"x\":296,\"y\":317},{\"x\":401,\"y\":317},{\"x\":401,\"y\":500},{\"x\":296,\"y\":486}],\"tableCellId\":0,\"word\":\"序号\",\"xec\":0,\"xsc\":0,\"yec\":1,\"ysc\":0},{\"pos\":[{\"x\":401,\"y\":317},{\"x\":711,\"y\":317},{\"x\":711,\"y\":503},{\"x\":401,\"y\":500}],\"tableCellId\":1,\"word\":\"考生号\",\"xec\":1,\"xsc\":1,\"yec\":1,\"ysc\":0}]
    """
    # 初始化最大值
    max_yec = 0
    max_xec = 0
    
    # 遍历数据，找出最大值

    for cell in cellInfos:
        max_yec = max(max_yec, cell['yec'])
        max_xec = max(max_xec, cell['xec'])
    
    # 行数为最大的 yec + 1（因为索引从0开始，而用户视角从1开始）
    # 列数为最大的 xec + 1（因为索引从0开始，而用户视角从1开始）
    row_count = max_yec + 1
    col_count = max_xec + 1
    
    return row_count, col_count

def aliocr_data_start_row(cellInfos):
    """
    获取ali云ocr表格识别后的json文件中的数据第一行。

    参数：
    - cellInfos：一个包含字典的列表，每个字典代表一个单元格的信息，包含'xsc'、'xec'、'ysc'、'yec'等键。

    返回值：表示数据第一行的行号。
    """
    # 初始化最大值变量，用于存储最大的 yec 和 ysc 的差值
    max_difference = 0

    # 遍历每个单元格的数据
    for item in cellInfos:
        # 计算当前单元格的 yec 和 ysc 的差值
        difference = item['yec'] - item['ysc']
        # 更新最大值
        if difference > max_difference:
            max_difference = difference

    return max_difference+1;

def aliocr_table_titles(cellInfos, title_row_toll):  
    """
    获取ali云ocr表格识别后的json文件中标题数据以及跨列的标题数据，包括word,xsc,xec,ysc,yec。返回的数据格式如下：
    [{'word': '序号', 'xsc': 0, 'xec': 0, 'ysc': 0, 'yec': 1},  {'word': '优惠加分', 'xsc': 19, 'xec': 19, 'ysc': 0, 'yec': 1}]
    title_list:表格的列名，可直接按列名导入到数据库。
    crossing_title_list：表格的跨列标头，如‘高考成绩’
    """
    title_list = []  # 用于存储标题项  
    crossing_title_list = []  # 用于存储跨列标题项  

    for cell in cellInfos:  
        # 获取单元格的 ysc 和 xsc, xec 值  
        ysc = cell['ysc']  
        xsc = cell['xsc']  
        xec = cell['xec']  

        # 检查 ysc 是否小于标题行数  
        if ysc < title_row_toll:  
            # 如果 xec - xsc == 0，添加到 title_list  
            if xec - xsc == 0:  
                title_list.append({  
                    'word': remove_all_whitespaces(cell['word']),  
                    'xsc': xsc,  
                    'xec': xec,  
                    'ysc': ysc,  
                    'yec': cell['yec']  
                })  
            else:  
                # 否则，添加到 crossing_title_list  
                crossing_title_list.append({  
                    'word': remove_all_whitespaces(cell['word']),  
                    'xsc': xsc,  
                    'xec': xec,  
                    'ysc': ysc,  
                    'yec': cell['yec']  
                })  
    # 按照 xsc 从小到大排序  
    title_list.sort(key=lambda x: x['xsc'])  
    crossing_title_list.sort(key=lambda x: x['xsc'])  

    if not crossing_title_list:
        crossing_title_list=aliocr_find_missing_crossing_cells(title_list);

    if title_row_toll>2:
        title_list=aliocr_correct_table_title(title_list,crossing_title_list)

    return title_list, crossing_title_list

def aliocr_filter_crossing_list(title_list, crossing_title_list):
    '''
    由于ocr识别的误差，表格标头在识别时会出现误差，导致本来是一列的元素识别成两列或多列，跨列元素多出来。如：
    2019-JX13.17-1/other\北京 (5).json
    [{'word': '政治|', 'xsc': 8, 'xec': 9, 'ysc': 0, 'yec': 1}, {'word': '民', 'xsc': 10, 'xec': 11, 'ysc': 0, 'yec': 1}, {'word': '高考成绩', 'xsc': 16, 'xec': 20, 'ysc': 0, 'yec': 0}, {'word': '', 'xsc': 16, 'xec': 17, 'ysc': 1, 'yec': 1}]
    解决思路是：删除xec-xsc大于3的跨列元素，将其重新添加到title_list
    '''
    new_crossing_title_list = []
    new_title_list = title_list.copy()
    for item in crossing_title_list:
        if item['xec'] - item['xsc'] > 3:
            new_crossing_title_list.append(item)
        else:
            new_title_list.append(item)
    return new_title_list, new_crossing_title_list

def aliocr_extract_key_from_cell(cell_infos,key):
    """
    从aliocr_table_titles(cellInfos, title_row_toll)的返回值中，获取表格标题列表。
    """
    # 初始化一个空列表来存储所有的word
    words_list = []
    
    # 遍历列表中的每个元素（每个元素都是一个字典）
    for cell in cell_infos:
        # 提取word值并添加到列表中
        words_list.append(cell[key])

    # 返回包含所有word的列表
    return words_list

def aliocr_create_rough_dataframe(cell_infos, title_list, data_start_row,col_index=0):  
    """
    将aliocr识别的json表格转化成df数据类型。
    """
    # 为cell_df的列设置空列表
    cell_tableCellId = []
    cell_yec = []
    cell_ysc = []
    cell_xec = []
    cell_xsc = []
    cell_num = []
    cell_word = []
    cell_axes = []
    st_num=''
    # 创建以 xsc 为索引，word 为值的字典，方便后续处理  
    title_dict = {title['xsc']: title['word'] for title in title_list}  

    # 计算 DataFrame 的行数，确保只包含需要的行  
    relevant_rows = [cell['ysc'] for cell in cell_infos if cell['ysc'] >= data_start_row]  
    max_row_index = max(relevant_rows) if relevant_rows else data_start_row - 1  

    # 创建一个空的 DataFrame，行数为 max_row_index + 1，列数为 title_list 的长度  
    df = pd.DataFrame(index=range(data_start_row, max_row_index + 1), columns=title_dict.values())  
    # df = pd.DataFrame(index=range(data_start_row, max_row_index + 1))  

    # 填充 DataFrame  
    for cell in cell_infos:  
        if cell['ysc'] >= data_start_row:  # 从指定的 ysc 行开始处理  
            xsc = cell['xsc']  
            word =  remove_all_whitespaces(cell['word'])

            # 检查该单元格的 xsc 是否在标题中  
            if xsc in title_dict: #以下直到284都在if之下  
                df.at[cell['ysc'], title_dict[xsc]] = word  # 填充 word 值 
                
                # 获取单元格的位置信息（字典形式），如果不存在则默认为空列表
                pos = cell.get('pos','')
                # 将临时坐标点列表添加到axes列表中，表示单元格的坐标信息
                cell_axes.append(pos)
                # 获取单元格的tableCellId，如果不存在则默认为0，转换为整数类型后添加到tableCellId列表中
                cell_tableCellId.append(int(cell.get('tableCellId', '-1')))
                word0 = cell.get('word', '')
                cell_word.append(word0)
                yec0 = cell.get('yec', '')
                cell_yec.append(yec0)
                ysc0 = cell.get('ysc', '')
                cell_ysc.append(ysc0)
                xec0 = cell.get('xec', '')
                cell_xec.append(xec0)
                xsc0 = cell.get('xsc', '')
                cell_xsc.append(xsc0)
                
                # 如果当前ysc0大于之前的最大ysc值
                if xsc0 ==0:
                    st_num = word0
                cell_num.append(st_num)

    # df_columns=split_list_by_length(title_list, df.shape[1])[0]
    # # 重新赋予列名
    # df.columns = df_columns

    # if df.columns[col_index]=='序号' or df.columns[col_index]=='顺序号':

    #     # 容错处理，检查给定列是否存在突变值
    #     bool_result,_,clean_result=check_for_discontinuities(df.iloc[:, col_index])

    #     # 如果存在突变值就纠错
    #     if not bool_result:
    #         correct_result=linear_interpolation(clean_result)
    #         df.iloc[:, 0]=correct_result.astype(int)

    if '序号' not in title_dict.values():
        df.insert(loc=0, column='序号', value=range(1, len(df) + 1))

    # 使用提取到的各种信息构建一个字典
    data = {
        'yec': cell_yec,
        'ysc': cell_ysc,
        'xec': cell_xec,
        'xsc': cell_xsc,
        'word': cell_word,
        'num': cell_num,
        'tableCellId': cell_tableCellId,
        'axes': cell_axes, #左上 XY 坐标、右上 XY 坐标、右下 XY 坐标、左下 XY 坐标。
    }
    # 使用字典创建一个DataFrame
    cell_df = pd.DataFrame(data)

    return df,cell_df


def aliocr_json_data_load(json_data):  
    """  
    读取aliocr识别的json数据，返回获取表格数据必要的json对象。  json_data是read_json_file的返回值。
    """  
    
    if "statusCode" not in json_data or json_data["statusCode"] != 200:  
        raise ValueError("OCR识别失败，状态码错误")  
    
    # 检查json_data是否是字典，并且statusCode是否为200  
    if not json_data or not isinstance(json_data, dict):  
        raise ValueError("未能读取JSON对象或JSON对象不是字典类型")  

    # 初始化返回值  
    result = {  
        'statusCode': None,  
        'cellInfos': [],  # 修改为列表以存储多个元素  
        'xCellSize': [],  # 修改为列表以存储多个元素  
        'yCellSize': [],  # 修改为列表以存储多个元素  
        'tableHeadTail': None,  
        'tableHead0': None,  
        'tableHead1': None,  
        'tablesCount': 0,  # 新增元素，表示表格数量  
        'content':None
    }  
    
    # 从JSON数据中获取 "Data" 字段的值，并进一步获取 "prism_tablesInfo" 和 "tableHeadTail" 部分  
    try:  
        json_headers_statusCode = json_data["statusCode"]  
        json_body_data = eval(json_data["body"]["Data"])  
        prism_tables_info = json_body_data["prism_tablesInfo"]  
        content=json_body_data["content"]
        
        # 遍历每个表格信息并提取相关数据  
        for table_info in prism_tables_info:  
            if table_info.get('tableId') != -1:  # 检查tableid是否不等于-1
                result['cellInfos'].append(table_info['cellInfos'])
                result['xCellSize'].append(table_info['xCellSize'])
                result['yCellSize'].append(table_info['yCellSize'])
                result['tablesCount'] += 1  # 只有当tableid不等于-1时，才增加表格数量

        tableHeadTail = json_body_data["tableHeadTail"]  
        tableHead0 = tableHeadTail[0]['head'][0] if tableHeadTail and len(tableHeadTail[0]['head']) > 0 else None  
        tableHead1 = tableHeadTail[0]['head'][1] if tableHeadTail and len(tableHeadTail[0]['head']) > 1 else None  

        # 更新结果字典  
        result.update({  
            'statusCode': json_headers_statusCode,  
            'tableHeadTail': tableHeadTail,  
            'tableHead0': tableHead0,  
            'tableHead1': tableHead1,
            'content':content 
        })  
    except Exception as e:  
        raise ValueError(f"处理JSON数据时发生错误: {e}")  
    
    return result

def aliocr_json_to_df_refine(json_pending_data,json_filepath):
    '''
    json_pending_data是函数aliocr_json_data_load的返回值。
    '''
    column_function_dict = {  
            3: partial(is_labels_series, values=['男', '女'], labels=('性别', '未识别')),  
        }
    
    data_start_row=aliocr_data_start_row(json_pending_data['cellInfos'][0])

    title_list, crossing_title_list=aliocr_table_titles(json_pending_data['cellInfos'][0],data_start_row)
    # print(title_list)
    # print(crossing_title_list)

    clean_df,cell_df=aliocr_create_rough_dataframe(json_pending_data['cellInfos'][0], title_list, data_start_row)

    contains_empty, empty_string_positions=contains_empty_string(aliocr_extract_key_from_cell(title_list,'word'))
    
    if contains_empty:
        clean_df,new_column_names=check_column_names(clean_df,column_function_dict)
        print(new_column_names)
        # 将new_column_names依次赋值给title_list的word
        for i in range(min(len(new_column_names), len(title_list))):
            title_list[i]['word'] = new_column_names[i]  

    if json_pending_data['tablesCount']==2:
        data_start_row1=0
        df1,cell_df=aliocr_create_rough_dataframe(json_pending_data['cellInfos'][1], title_list, data_start_row1)
        clean_df,new_column_names=check_column_names(df1,column_function_dict)
    

    if crossing_title_list: # 跨列的标题列表，默认只有一个跨列标题，所以只处理了一个跨列标题
        # 从crossing_title_list中获取xsc和xec的值
        xsc, xec = crossing_title_list[0]['xsc'], crossing_title_list[0]['xec']
        # 根据xsc值从title_list中获取word值
        merged_columns = [item['word'] for item in title_list if item['xsc'] >= xsc and item['xsc'] <= xec]
        crossing_column_name='高考成绩'
        clean_df = merge_columns_to_dict_series(clean_df, merged_columns,crossing_column_name )

    # 定义正则表达式模式,提取著录信息
    year_pattern = r"((?:19|20)\d{2})(?:年|普|.*名册|\-)" 
    province_pattern=r'(北京|天津|上海|重庆|河北|山西|内蒙古|辽宁|吉林|黑龙江|江苏|浙江|安徽|福建|江西|山东|河南|湖北|湖南|广东|海南|四川|贵州|云南|陕西|甘肃|青海|台湾|广西|西藏|宁夏|新疆|香港|澳门|华侨港澳台|港澳台)(?:学生|.*名册|壮族|回族|维吾尔|维尔|省|自治区|市)'
    admission_pattern=r'(\S+科\S+批|\S批|批次\S+|\S+阶段|\S+类\S+|层次\S+|\S+专项|\S+计划|\科类S+)'
    year = extract_str_by_pattern(json_filepath,year_pattern)
    province = extract_str_by_pattern(json_filepath,province_pattern)
    admission_patch=extract_str_by_pattern(json_pending_data['content'],admission_pattern)
    if province is None or province=='':
        admission_patch = extract_str_from_pattern_in_json_list(json_pending_data['tableHeadTail'], admission_pattern)  
    if province is None or province=='':
        province=extract_str_from_pattern_in_json_list(json_pending_data['tableHeadTail'], province_pattern)
    if province is None or province=='':
        province=extract_str_by_pattern(json_pending_data['content'],province_pattern)
    if year is None or year=='':
        year=extract_str_from_pattern_in_json_list(json_pending_data['tableHeadTail'],year_pattern)
    if year is None or year=='':
        year=extract_str_by_pattern(json_pending_data['content'],year_pattern)

    clean_df['year']=year
    clean_df['province']=province
    clean_df['admissionBatch']=admission_patch
    clean_df['created_at']=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # 在DataFrame中添加xCellSize和yCellSize列
    cell_df['province'] = province
    cell_df['year'] = year
    cell_df['admissionBatch'] = admission_patch
    cell_df['created_at'] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    content_df_data = {'province': province, 'year': year, 'created_at': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
    content_df = pd.DataFrame(data=[content_df_data])

    return clean_df,cell_df,content_df #clean_df是数据表，cell_df是起辅助作用的单元格信息表

def aliocr_image_to_json(filepath,fp_setting,other_dir='other'):
    '''
    fp_setting是配置文件的路径，里面有ali的应用id和secret，名称分别是OCR.ALIBABA_CLOUD_ACCESS_KEY_SECRET和OCR.ALIBABA_CLOUD_ACCESS_KEY_ID
    '''
    filepath = os.path.abspath(filepath) # file的绝对路径
    fn = os.path.basename(filepath)
    filename_without_extention=os.path.splitext(fn)[0] # file不带扩展名的名称
    if other_dir=='other':
        other_path=check_create_folder(os.path.join(os.path.dirname(filepath),other_dir)) # file所在文件夹下的other文件夹
    else:
        other_path=other_dir

    json_file_name=filename_without_extention+"-ali-json.json" # json文件夹下的json文件名

    full_jsonfile_path=os.path.join(other_path,json_file_name) # json文件完整路径

    if check_file_existence(full_jsonfile_path)==False:
        # 使用 Pillow 打开图像文件
        PILimg = Image.open(filepath)

        if check_image_resolution(PILimg, 8100, 8100)==False:
            print("开始改变分辨率")
            PILimg_new=resize_image_resolution(PILimg, 8100)
            # 将图像对象转换为文件对象
            img = io.BytesIO()
            PILimg_new.save(img, format='JPEG')
            img.seek(0)
            print("已改变分辨率")
        else:
            img = open(filepath, 'rb')
            
            if(get_image_file_size(img)>10200):
                print("开始改变大小")
                img=resize_image_to_target_size(img)
            print("已改变大小")
        print("开始上传阿里识别")
        json_data=json.loads(Aliyun_table_Ocr.main(img,fp_setting))

        # 将 JSON 数据存入文件
        with open(full_jsonfile_path, "w",encoding="utf-8") as file:

            json.dump(json_data, file, ensure_ascii=False,indent=4) # 将 JSON 格式的字符串写入 JSON 格式的文件
        printLog("INFORMATION:", f'{fn} 已解析为 json 文件，存储在 {full_jsonfile_path}')
        return json_data

def aliocr_correct_table_title(tabel_title_list, corossing_title_list):
    """
    这个函数主要用于处理表格标题列表，进行数据合并和更新操作。

    参数：
    - tabel_title_list：一个包含字典的列表，每个字典代表一个表格标题项，包含'xsc'、'xec'、'ysc'、'yec'和'word'等键。
    - corossing_title_list：另一个包含字典的列表，用于特定条件判断。

    返回值：处理后的表格标题列表，其中进行了合并和更新操作。
    """
    # 分组数据
    grouped_data = {}
    for item in tabel_title_list:
        # 使用 (item['xsc'], item['xec']) 作为键对数据进行分组
        key = (item['xsc'], item['xec'])
        if key not in grouped_data:
            grouped_data[key] = []
        grouped_data[key].append(item)

    # 处理每组数据
    for key, items in grouped_data.items():
        if len(items) > 1:
            # 找到 ysc 最小的元素
            min_item = min(items, key=lambda x: x['ysc'])
            # 合并 word 值
            merged_word = ''.join(item['word'] for item in items)
            # 更新最小 ysc 元素的 word 和 yec
            min_item['word'] = merged_word
            min_item['yec'] = 1
            # 将其他元素标记为删除
            for item in items:
                if item!= min_item:
                    item['_delete'] = True

    # 更新 ysc 和 yec 的值，如果 xsc 在 data2 范围内
    for item in tabel_title_list:
        if corossing_title_list[0]['xsc'] <= item['xsc'] <= corossing_title_list[0]['xec']:
            item['ysc'] = 1
            item['yec'] = 1

    # 移除标记为删除的元素，并返回结果
    return [item for item in tabel_title_list if '_delete' not in item]

def aliocr_find_missing_crossing_cells(title_list):  
    '''
    表格标头格式：[{'word': '高考成绩', 'xsc': 14, 'xec': 25, 'ysc': 0, 'yec': 0},......]

    将表格标头缺失的跨列元素找回来，默认word值为‘高考成绩’。
    '''
    # 创建一个二维列表来表示单元格  
    max_rows = max(item['yec'] for item in title_list) + 1  # 找到最大行数  
    max_cols = max(item['xec'] for item in title_list) + 1  # 找到最大列数  
    
    # 初始化一个网格来表示已占用的单元格  
    grid = [[False] * max_cols for _ in range(max_rows)]  
    
    # 标记已占用的单元格  
    for item in title_list:  
        for row in range(item['ysc'], item['yec'] + 1):  
            for col in range(item['xsc'], item['xec'] + 1):  
                grid[row][col] = True  
                
    # 找到缺失的单元格  
    missing_cells = []  
    
    for row in range(max_rows):  
        start_col = None  
        
        for col in range(max_cols):  
            if not grid[row][col]:  # 如果该单元格为空  
                if start_col is None:  
                    start_col = col  # 开始新的空白列  
            else:  
                if start_col is not None:  
                    # 如果之前有空集合，记录缺失的区域  
                    missing_cells.append({  
                        'word': '高考成绩',  
                        'xsc': start_col,  
                        'xec': col - 1,  
                        'ysc': row,  
                        'yec': row  
                    })  
                    start_col = None  # 重置开始列  

        # 检查是否在行的末尾有空白区域  
        if start_col is not None:  
            missing_cells.append({  
                'word': '高考成绩',  
                'xsc': start_col,  
                'xec': max_cols - 1,  
                'ysc': row,  
                'yec': row  
            })  

    # 合并所有缺失单元格，取最小xsc和最大xec  
    if missing_cells:  
        final_missing_cell = {  
            'word': '高考成绩',  
            'xsc': min(cell['xsc'] for cell in missing_cells),  
            'xec': max(cell['xec'] for cell in missing_cells),  
            'ysc': missing_cells[0]['ysc'],  
            'yec': missing_cells[0]['yec']  
        }  
        return [final_missing_cell]  

    return []  

def ocr_build_other_filepath(file_path, ocr_model='ali'):
    '''
    根据文件路径生成ocr识别json路径，table、cell、content、table-content路径。
    '''
    try:
        # 获取文件名（带扩展名）
        filename = os.path.basename(file_path)

        # 分离文件名和扩展名
        filename_without_extension, file_extension = os.path.splitext(filename)

        # 定义基础目录
        base_dir = os.path.dirname(file_path)
        other_dir = base_dir  # 默认为基础目录
        file_path=file_path.replace('\\', '/')

        # 如果文件扩展名是 json，调整目录和文件名处理
        if file_extension.lower() == '.json':
            ori_full_name=find_ori_files_from_json(file_path)
            ori_filename = ori_full_name[0]
            filename_without_extension,_= os.path.splitext(ori_full_name[1])
        elif file_extension.lower() in ['.jpg','.jpeg','.tif','.tiff']:
            # 定义其他文件的目录
            other_dir = os.path.join(base_dir, 'other')
            ori_filename = file_path

        else:
            print('ocr_build_other_filepath 不支持的文件格式。')
            return None
        if ocr_model == 'paddle':
            filename_without_extension = f'{filename_without_extension}-paddle'
        elif ocr_model=='ali':
            filename_without_extension = f'{filename_without_extension}-ali'
        else:
            print('ocr_build_other_filepath 不支持的ocr Model。')
            return None
        # 构建其他文件路径
        json_filename = os.path.join(other_dir, f'{filename_without_extension}-json.json')
        table_filename = os.path.join(other_dir, f'{filename_without_extension}-table.xlsx')
        cell_filename = os.path.join(other_dir, f'{filename_without_extension}-cell.xlsx')
        content_filename = os.path.join(other_dir, f'{filename_without_extension}-content.docx')
        table_content_filename = os.path.join(other_dir, f'{filename_without_extension}-table_content.xlsx')
        log_filename = os.path.join(other_dir, 'aliocr-log.xlsx')
        if ocr_model == 'paddle':
            log_filename = os.path.join(other_dir, 'paddleocr-log.xlsx')

        # 返回生成的文件路径
        return ori_filename, json_filename, table_filename, cell_filename, content_filename, log_filename, table_content_filename
    except Exception as e:
        print(f"ocr_build_other_filepath-An error occurred: {e}")
        return None
    
def aliocr_json_to_file(aliocr_json_data,filenames,json_filepath,table_flag=True,cell_flag=True,content_flag=True,content_type='gklqmc'):
    '''
    将json数据转化为table/cell/content文件。table可以直接查询或存入数据库，cell是表格单元格详细信息，可进一步对表格进行自动处理，如遮盖选定行或列，content就是表格中所有信息文本word文件，可用来进行粗略查询。
    '''
    printLog(f"aliocr_json_to_file 开始转换 {filenames[1]} ")
    try:
        json_pending_data=aliocr_json_data_load(aliocr_json_data)
        
        table_df,cell_df,content_df=aliocr_json_to_df_refine(json_pending_data,json_filepath=json_filepath)
        
        # 定义正则表达式来匹配特定格式（1950 - JX11.44 - 10形式）的字符，用于从文件路径中获取archiveCode
        archive_file_pattern = r'\d{4}-[A-Za-z0-9]{2}\d{2}\.[A-Za-z0-9]{2}-\d+'
        # 使用正则表达式在filepath中进行匹配，获取匹配到的archiveCode
        archiveCode = re.search(archive_file_pattern, filenames[0]).group(0) if re.search(archive_file_pattern, filenames[0]) else ""
        table_df['archiveCode']=archiveCode
        cell_df['archiveCode'] = archiveCode
        archiveClass=extract_archive_class(archiveCode)
        table_df['archiveClass']=archiveClass

        relative_path=split_path(filenames[0], 'da')
        linux_filepath= convert_path_to_system_style(relative_path,'linux')
        table_df['filePath'] = linux_filepath
        # 在DataFrame的第一列插入文件路径列
        cell_df.insert(loc=0, column='filePath', value=linux_filepath)

        json_log_flag=0
        table_log_flag=0
        cell_log_flag=0
        table_content_log_flag=0
        content_log_flag=0
        db_log_flag=0

    except Exception as e:
        printLog(f"aliocr_json_to_file转换 table_content 发生错误：{e}")


    try:
        if content_flag==True:
            content_df['type']=content_type
            content_df['content']=json_pending_data['content']
            content_df['fullPath_link']=convert_path_to_system_style(filenames[0],'linux')
            content_df['filePath']=linux_filepath
            content_df['archiveCode']=archiveCode
            content_df['archiveClass']=archiveClass
            content_df.to_excel(filenames[6], index = False)

            printLog(f"aliocr_json_to_file处理完成table_content： {filenames[6]} ")
            table_content_log_flag=1

        if content_flag==True:

            text_to_docx(filenames[4],json_pending_data['content'],head='')

            printLog(f"aliocr_json_to_file处理完成content： {filenames[4]} ")
            content_log_flag=1

    except Exception as e:
        printLog(f"aliocr_json_to_file转换 content 发生错误：{e}")

    try:
        if table_flag==True:
            table_df.to_excel(filenames[2], index = False)
            printLog(f"aliocr_json_to_file处理完成table： {filenames[2]} ")
            table_log_flag=1
    except Exception as e:
        printLog(f"aliocr_json_to_file转换 table 发生错误：{e}")

    try:

        if cell_flag==True:
            cell_df.to_excel(filenames[3], index = False)
            printLog(f"aliocr_json_to_file处理完成cell： {filenames[3]} ")
            cell_log_flag=1

    except Exception as e:
        printLog(f"aliocr_json_to_file转换 cell 发生错误：{e}")

    printLog(f"aliocr_json_to_file 处理完成. ")
    log_at=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    check_json=check_file_existence(filenames[1])

    if check_json:
        json_log_flag=1


    return [convert_path_to_system_style(filenames[0],'linux'), json_log_flag, table_log_flag, cell_log_flag, table_content_log_flag,content_log_flag,db_log_flag,filenames[1],filenames[2], filenames[3],filenames[6],filenames[4],log_at]

def aliocr_log(log_dict):
    '''
    log_dict格式是{源文件名1:log_df_row,源文件名2:log_df_row}，其中log_df_row是函数paddleocr_json_to_file的返回值，是一个源文件的log记录。
    '''
    log_fields = ['file_path', 'json_flag', 'table_flag', 'cell_flag','table_content_log_flag', 'content_flag','db_flag','json_path','table_path','cell_path','table_content_path','content_path','log_at']
    
    # 遍历字典
    for fp, values in log_dict.items():
    
        # 将二维数组转换为DataFrame，并指定列名
        if not values:  # 检查是否有数据
            continue
        new_data = pd.DataFrame(values, columns=log_fields)
    
        # 检查文件是否存在
        if check_file_existence(fp):
            # 文件存在，读取文件
            log_df = pd.read_excel(fp)
            # 删除列名后缀有 '_Link' 的列
            log_df = log_df.loc[:, ~log_df.columns.str.endswith('_Link')]
            # 检查待插入的数据是否已经存在
            for index, row in new_data.iterrows():
                if not log_df[log_df['file_path'] == row['file_path']].empty:
                    printLog(f"aliocr_log 数据已存在，将进行覆盖：{row['file_path']}")
                    # 获取匹配的行的索引
                    match_index = log_df[log_df['file_path'] == row['file_path']].index[0]
                    # 更新匹配的行
                    log_df.loc[match_index, :] = row
                else:
                    log_df = pd.concat([log_df, row.to_frame().T], ignore_index=True)
            
            delete_file(fp)
        else:
            # 文件不存在，使用新的DataFrame
            log_df = new_data
        
        path_columns=['file_path','json_path','table_path','cell_path','table_content_path','content_path']
        save_df_with_links(log_df, path_columns, fp)
        printLog(f"log 文件已生成：{fp} ")

def aliocr_json_to_file_with_log(file_paths,table_flag=True,cell_flag=True,content_flag=True):
    '''
    log_dict格式是{源文件名1:log_df_row,源文件名2:log_df_row}，其中log_df_row是函数paddleocr_json_to_file的返回值，是一个源文件的log记录。
    '''
    log_dict={}
    process_num=1
    total_files=len(file_paths)
    printLog(f"文件总数：{total_files}")
    for file_path in file_paths:
        printLog(f"aliocr_json_to_file_with_log 开始处理：")

        try:
            aliocr_json_data=read_json_file(file_path)
            filenames=ocr_build_other_filepath(file_path)
            log_df_row=aliocr_json_to_file(aliocr_json_data,filenames,file_path,table_flag,cell_flag,content_flag)

            log_path=filenames[5]
            # 检查键是否存在，如果不存在则添加键及其初始值
            if log_path not in log_dict:
                log_dict[log_path] = [log_df_row]
            else:
                # 向键对应的二维数组添加新的子列表
                log_dict[log_path].append(log_df_row)
            printLog(f"进度: {process_num}/{total_files}")
            process_num+=1
        except Exception as e:
            printLog(f"aliocr_json_to_file_with_log 发生错误：{e}")
            process_num+=1

    aliocr_log(log_dict)  

def aliocr_extract_json_columns_to_excel_with_errors(total_files,excel_path,files_paths):
    '''
    从表格json文件中抽取列名，为下一步校正列名导入数据库做准备。
    '''
    processed_files = 0
    df = pd.DataFrame()
    df_errors=pd.DataFrame()
    for file_path in files_paths:
        try:
            df_temp= pd.DataFrame()
            from_file=''
            result=[]
            data_start_row=''
            title_list=[]
            crossing_title_list=[]
            words=[]
            xscs=[]
            from_file=find_ori_files_from_json(file_path)
            aliocr_json_data=read_json_file(file_path)
            result=aliocr_json_data_load(aliocr_json_data)
            data_start_row=aliocr_data_start_row(result['cellInfos'][0])
            title_list, crossing_title_list=aliocr_table_titles(result['cellInfos'][0],data_start_row)

            # 定义正则表达式模式,提取著录信息
            year_pattern = r"((?:19|20)\d{2})(?:年|普|.*名册|\-)" 
            province_pattern=r'(北京|天津|上海|重庆|河北|山西|内蒙古|辽宁|吉林|黑龙江|江苏|浙江|安徽|福建|江西|山东|河南|湖北|湖南|广东|海南|四川|贵州|云南|陕西|甘肃|青海|台湾|广西|西藏|宁夏|新疆|香港|澳门|华侨港澳台|港澳台)(?:学生|.*名册|壮族|回族|维吾尔|维尔|省|自治区|市)'
            admission_pattern=r'(\S+科\S+批|\S批|批次\S+|\S+阶段|\S+类\S+|层次\S+|\S+专项|\S+计划|\科类S+)'
            year = extract_str_by_pattern(file_path,year_pattern)
            province = extract_str_by_pattern(file_path,province_pattern)
            admission_patch=extract_str_by_pattern(result['content'],admission_pattern)
            if province is None or province=='':
                admission_patch = extract_str_from_pattern_in_json_list(result['tableHeadTail'], admission_pattern)  
            if province is None or province=='':
                province=extract_str_from_pattern_in_json_list(result['tableHeadTail'], province_pattern)
            if province is None or province=='':
                province=extract_str_by_pattern(result['content'],province_pattern)
            if year is None or year=='':
                year=extract_str_from_pattern_in_json_list(result['tableHeadTail'],year_pattern)
            if year is None or year=='':
                year=extract_str_by_pattern(result['content'],year_pattern)

            # 定义正则表达式来匹配特定格式（1950 - JX11.44 - 10形式）的字符，用于从文件路径中获取archiveCode
            archive_file_pattern = r'\d{4}-[A-Za-z0-9]{2}\d{2}\.[A-Za-z0-9]{2}-\d+'
            # 使用正则表达式在filepath中进行匹配，获取匹配到的archiveCode
            archiveCode = re.search(archive_file_pattern, file_path).group(0) if re.search(archive_file_pattern,file_path) else ""

            words = [item['word'] for item in title_list]
            xscs = [str(item['xsc']) for item in title_list]

            # 创建一个新列，初值为 0
            cross_title_column = [0] * len(xscs)
            try:
                cross_title_xsc = crossing_title_list[0]['xsc']
                cross_title_xec = crossing_title_list[0]['xec']
                for i in range(len(xscs)):
                    if cross_title_xsc <= int(xscs[i]) <= cross_title_xec:
                        cross_title_column[i] = 1
            except IndexError:
                # 如果出现列表索引越界错误，直接跳过这段逻辑继续执行后续代码
                pass

            df_temp = pd.DataFrame({'word': words, 'xsc': xscs,'cross_title':cross_title_column})
            df_temp['year']=year
            df_temp['province']=province
            df_temp['admission_patch']=admission_patch
            df_temp['archiveCode']=archiveCode
            df_temp['file_path']=file_path
            df_temp['from_file']=from_file[0]
            # 使用 concat 函数将 df_temp 垂直堆叠添加到 df
            df = pd.concat([df, df_temp], ignore_index=True)
            processed_files += 1
            print(f"Progress: {processed_files}/{total_files}")
        except Exception as e:
            df_errors_temp=pd.DataFrame()
            # 创建一个包含所有信息的字典
            if from_file:  # 检查列表是否非空
                df_errors_temp = pd.DataFrame({'file_path': [file_path], 'link': file_path,'from_file': from_file[0],'from_link':from_file[0],'errors':str(e)})
            else:
                df_errors_temp = pd.DataFrame({'file_path': [file_path], 'link': file_path,'from_file': None,'from_link':None,'errors':str(e)})
    
            df_errors = pd.concat([df_errors, df_errors_temp], ignore_index=True)
            processed_files += 1
            print(f"Progress: {processed_files}/{total_files}")
            printLog(f"aliocr_extract_json_columns_to_excel_with_errors出现错误：{e}")
    df.insert(1, 'actual_value', np.nan)
    df.insert(0, 'db_key', np.nan)
    path_columns=['file_path','from_file']
    
    save_df_with_links(df,path_columns,excel_path)
    
    if not df_errors.empty:
        # 创建一个 Excel 工作簿
        # wb = Workbook()
        # 加载现有的工作簿
        wb = load_workbook(excel_path)
        # 检查工作表是否存在
        if 'errors' not in wb.sheetnames:
            # 创建一个新的工作表并命名为 'errors'
            errors_ws = wb.create_sheet(title='errors')
        else:
            errors_ws = wb['errors']  # 如果存在则获取现有工作表 
        # 将新创建的工作表设置为当前活动的工作表
        wb.active = errors_ws
        # 将 DataFrame 转换为 Excel 行
        for r_idx, row in enumerate(dataframe_to_rows(df_errors, index=False, header=True), 1):
            # 遍历行中的每个单元格
            for c_idx, value in enumerate(row, 1):
                # 将值写入单元格
                errors_ws.cell(row=r_idx, column=c_idx, value=value)
                # 如果是文件路径列（假设是第四列），并且不是标题行，则添加超链接
                if (c_idx == 2 or c_idx == 4) and r_idx > 1:  # r_idx > 1 确保跳过标题行
                    errors_ws.cell(row=r_idx, column=c_idx).hyperlink = str(value)
                    errors_ws.cell(row=r_idx, column=c_idx).value = 'Open File'
                    # 保存工作簿
        wb.save(excel_path)

def paddleocr_extract_word_from_jsondata(json_data):
    '''
    将paddleocr识别的json数据转化为字符串，json_data是paddle识别后的原始json对象。
    '''

    extracted_texts=[]
    for h in json_data:
        for i in h:
            extracted_texts.append(str(i[1][0]))
    result=" ".join(extracted_texts) 
    return result

def paddle_image_to_json(filepath,fp_setting=None,other_dir='other'):
    """
    用 PaddleOCR 将 image 识别为 json 数据。
    fp_setting是配置文件路径，暂时不用。

    参数：
    - filepath (str): 要进行 OCR 识别的图像文件路径。
    - json_file_path (str, optional): 可选参数，保存结果的 JSON 文件路径。如果为 None，则不保存到文件。
    - gpu (bool): 是否使用 GPU 进行计算。
    - use_angle_cls (bool): 是否使用角度分类器。
    - lang (str): 识别的语言，默认为'ch'（中文）。
    - show_log (bool): 是否显示日志信息。
    - other_dir: 保存路径

    返回：
    - result (list): OCR 识别的结果列表。

    功能描述：
    1. 根据 gpu 参数设置设备为 GPU 或 CPU。
    2. 初始化 PaddleOCR 对象，设置相关参数。
    3. 进行 OCR 识别。
    4. 如果提供了 json_file_path，则将结果保存到文件中。
    """
    filepath = os.path.abspath(filepath) # file的绝对路径
    fn = os.path.basename(filepath)
    filename_without_extention=os.path.splitext(fn)[0] # file不带扩展名的名称
    if other_dir=='other':
        other_path=check_create_folder(os.path.join(os.path.dirname(filepath),other_dir)) # file所在文件夹下的other文件夹
    else:
        other_path=other_dir

    json_file_name=filename_without_extention+"-paddle-json.json" # json文件夹下的json文件名

    full_jsonfile_path=os.path.join(other_path,json_file_name) # json文件完整路径

    gpu=False
    use_angle_cls=True
    lang='ch'
    show_log=False
    try:
        os.environ["KMP_DUPLICATE_LIB_OK"]="TRUE" #加入这段代码可以解决使用paddleocr时内核重启的问题
        # 学习网址 https://paddlepaddle.github.io/PaddleOCR/ppocr/quick_start.html#221
        # 或者可以尝试conda uninstall jupyter tornado  然后重新安装conda install jupyter 

        # 这里的环境变量 KMP_DUPLICATE_LIB_OK 是特定于Intel的Math Kernel Library (MKL) 的。MKL是用于科学计算的一系列例程，它被许多数据处理和机器学习库所使用，包括NumPy和一些深度学习框架。

        # 在某些情况下，如果你的系统中安装了多个版本的MKL或者其他与MKL相关的库，可能会出现动态链接库（DLL）冲突的问题。这可能会导致程序在运行时抛出错误，比如“已加载重复的MKL库”之类的消息。

        # 通过设置环境变量 KMP_DUPLICATE_LIB_OK 为 TRUE，你可以告诉程序忽略这些冲突，允许加载多个MKL库的副本。这通常是一种临时解决方案，用于在不解决底层库冲突的情况下让程序运行。
        if check_file_existence(full_jsonfile_path):
            print(f'{filepath}已存在。')
            return []
        if gpu:
            # 设置使用 GPU
            try:
                paddle.set_device('gpu')
            except paddle.fluid.core.EnforceNotMetError as e:
                # 如果设置 GPU 失败，输出错误信息并改为使用 CPU
                print(f"无法使用 GPU，错误信息：{e}. 改为使用 CPU。")
                paddle.set_device('cpu')
        
        # 初始化模型，使用指定语言识别
        ocr = PaddleOCR(use_angle_cls=use_angle_cls, lang=lang, show_log=show_log)

        # 进行 OCR 识别
        json_data = ocr.ocr(filepath, cls=True)
        
        if full_jsonfile_path is not None:
            # 可以选择将结果保存到文件
            try:
                with open(full_jsonfile_path, 'w', encoding='utf-8') as f:
                    f.write(json.dumps(json_data, ensure_ascii=False))
                    print(f"paddle_image_to_json - json文件保存到：{full_jsonfile_path}")
            except IOError as e:
                print(f"保存结果到文件时出现错误：{e}")
        return json_data
        
    except Exception as e:
        print(f"发生未知错误：{e}")
        return []
    
def paddleocr_json_to_df_refine(json_pending_data,json_filepath):
    '''
    json_pending_data是函数paddle_image_to_json的返回值。
    '''

    clean_df=pd.DataFrame()
    cell_df=pd.DataFrame()

    words=paddleocr_extract_word_from_jsondata(json_pending_data)

    # 定义正则表达式模式,提取著录信息
    year_pattern = r"((?:19|20)\d{2})(?:年|普|.*名册|\-)" 
    province_pattern=r'(北京|天津|上海|重庆|河北|山西|内蒙古|辽宁|吉林|黑龙江|江苏|浙江|安徽|福建|江西|山东|河南|湖北|湖南|广东|海南|四川|贵州|云南|陕西|甘肃|青海|台湾|广西|西藏|宁夏|新疆|香港|澳门|华侨港澳台|港澳台)(?:学生|.*名册|壮族|回族|维吾尔|维尔|省|自治区|市)'
    admission_pattern=r'(\S+科\S+批|\S批|批次\S+|\S+阶段|\S+类\S+|层次\S+|\S+专项|\S+计划|\科类S+)'
    province=extract_str_by_pattern(json_filepath,province_pattern)
    year=extract_str_by_pattern(words,year_pattern)
    admission_batch = extract_str_by_pattern(words, admission_pattern)  

    if province is None or province=='':
        province=extract_str_by_pattern(json_filepath,province_pattern)

    if year is None or year=='':
        year=extract_str_by_pattern(json_filepath,year_pattern)

    content_df_data = {'province': province, 'year': year, 'admission_batch':admission_batch,'content':words,'created_at': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
    content_df = pd.DataFrame(data=[content_df_data])

    return clean_df,cell_df,content_df

def paddleocr_json_to_file(paddleocr_json_data,filenames,json_filepath,table_flag=False,cell_flag=False,content_flag=True,content_type='gklqmc'):
    '''
    将json数据转化为table/cell/content文件。table可以直接查询或存入数据库，cell是表格单元格详细信息，可进一步对表格进行自动处理，如遮盖选定行或列，content就是表格中所有信息文本word文件，可用来进行粗略查询。
    log_fields = ['file_path', 'json_flag', 'table_flag', 'cell_flag','table_content_log_flag', 'content_flag','db_flag','json_path','table_path','cell_path','table_content_path','content_path','log_at']
    '''
    printLog(f"paddleocr_json_to_file 开始转换 {filenames[1]} ")


    table_df,cell_df,content_df=paddleocr_json_to_df_refine(paddleocr_json_data,json_filepath)

    # 定义正则表达式来匹配特定格式（1950 - JX11.44 - 10形式）的字符，用于从文件路径中获取archiveCode
    archive_code_pattern = r'\d{4}-[A-Za-z0-9]{2}\d{2}\.[A-Za-z0-9]{2}-\d+'
    # 使用正则表达式在filepath中进行匹配，获取匹配到的archiveCode
    archiveCode = re.search(archive_code_pattern, filenames[0]).group(0) if re.search(archive_code_pattern, filenames[0]) else ""
    table_df['archiveCode']=archiveCode
    cell_df['archiveCode'] = archiveCode
    archiveClass=extract_archive_class(archiveCode)
    table_df['archiveClass']=archiveClass

    relative_path=split_path(filenames[0], 'da')
    linux_filepath= convert_path_to_system_style(relative_path,'linux')
    table_df['filePath'] = linux_filepath
    # 在DataFrame的第一列插入文件路径列
    cell_df.insert(loc=0, column='filePath', value=linux_filepath)

    json_log_flag=0
    table_log_flag=0
    cell_log_flag=0
    table_content_log_flag=0
    content_log_flag=0
    db_log_flag=0

    try:

        if content_flag==True:
            content_df['type']=content_type
            content_df['filePath']=linux_filepath
            content_df['fullPath_link']=filenames[0]
            content_df['archiveCode']=archiveCode
            content_df['archiveClass']=archiveClass
            content_df.to_excel(filenames[6], index = False)

            printLog(f"paddleocr_json_to_file处理完成： {filenames[6]} ")
            table_content_log_flag=1

    except Exception as e:
        printLog(f"paddleocr_json_to_file转换 table_content 发生错误：{e}")

    try:

        if content_flag==True:

            text_to_docx(filenames[4],str(content_df['content'].iloc[0]),head='')

            printLog(f"paddleocr_json_to_file转换 content： {filenames[4]} ")
            content_log_flag=1

    except Exception as e:
        printLog(f"paddleocr_json_to_file转换 content 发生错误：{e}")

    try:

        if table_flag==True:
            table_df.to_excel(filenames[2], index = False)
            printLog(f"paddleocr_json_to_file处理完成： {filenames[2]} ")
            table_log_flag=1
    except Exception as e:
        printLog(f"paddleocr_json_to_file转换 table 发生错误：{e}")

    try:

        if cell_flag==True:
            cell_df.to_excel(filenames[3], index = False)
            printLog(f"paddleocr_json_to_file处理完成： {filenames[3]} ")
            cell_log_flag=1
    except Exception as e:
        printLog(f"转换 cell 发生错误：{e}")
    check_json=check_file_existence(filenames[1])

    if check_json:
        json_log_flag=1
    printLog(f"paddleocr_json_to_file 处理完成. ")
    log_at=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    return [convert_path_to_system_style(filenames[0],'linux'), json_log_flag, table_log_flag, cell_log_flag, table_content_log_flag,content_log_flag,db_log_flag,filenames[1],filenames[2], filenames[3],filenames[6],filenames[4],log_at]

def paddleocr_json_to_file_with_log(file_paths,table_flag=False,cell_flag=False,content_flag=True):
    '''
    file_paths是json文件路径列表。
    log_dict格式是{源文件名1:log_df_row,源文件名2:log_df_row}，其中log_df_row是函数paddleocr_json_to_file的返回值，是一个源文件的log记录。
    '''
    log_dict={}
    process_num=1
    total_files=len(file_paths)
    printLog(f"文件总数：{total_files}")
    for file_path in file_paths:
        printLog(f"paddleocr_json_to_file_with_log 开始处理：")
        try:
            ocr_json_data=read_json_file(file_path)
            filenames=ocr_build_other_filepath(file_path,'paddle')

            log_df_row=paddleocr_json_to_file(ocr_json_data,filenames,file_path,table_flag,cell_flag,content_flag)
            
            log_path=filenames[5]
            # 检查键是否存在，如果不存在则添加键及其初始值
            if log_path not in log_dict:
                log_dict[log_path] = [log_df_row]
            else:
                # 向键对应的二维数组添加新的子列表
                log_dict[log_path].append(log_df_row)
            printLog(f"进度: {process_num}/{total_files}")
            process_num+=1
        except Exception as e:
            printLog(f"paddleocr_json_to_file_with_log 发生错误：{e}")
            process_num+=1

    paddleocr_log(log_dict)    

def paddleocr_log(log_dict):
    '''
    log_dict格式是{源文件名1:log_df_row,源文件名2:log_df_row}，其中log_df_row是函数paddleocr_json_to_file的返回值，是一个源文件的log记录。
    '''
    log_fields = ['file_path', 'json_flag', 'table_flag', 'cell_flag','table_content_log_flag', 'content_flag','db_flag','json_path','table_path','cell_path','table_content_path','content_path','log_at']
    
    # 遍历字典
    for fp, values in log_dict.items():
    
        # 将二维数组转换为DataFrame，并指定列名
        if not values:  # 检查是否有数据
            continue
        new_data = pd.DataFrame(values, columns=log_fields)
    
        # 检查文件是否存在
        if check_file_existence(fp):
            # 文件存在，读取文件
            log_df = pd.read_excel(fp)
            # 删除列名后缀有 '_Link' 的列
            log_df = log_df.loc[:, ~log_df.columns.str.endswith('_Link')]
            # 检查待插入的数据是否已经存在
            for index, row in new_data.iterrows():
                if not log_df[log_df['file_path'] == row['file_path']].empty:
                    printLog(f"数据已存在，将进行覆盖：{row['file_path']}")
                    # 获取匹配的行的索引
                    match_index = log_df[log_df['file_path'] == row['file_path']].index[0]
                    # 更新匹配的行
                    log_df.loc[match_index, :] = row
                else:
                    log_df = pd.concat([log_df, row.to_frame().T], ignore_index=True)
            
            delete_file(fp)
        else:
            # 文件不存在，使用新的DataFrame
            log_df = new_data
        
        path_columns=['file_path','json_path','table_path','cell_path','table_content_path','content_path']
        save_df_with_links(log_df, path_columns, fp)
        printLog(f"log 文件已生成：{fp} ")

def display_rectangle_in_image(json_data, image_path, save_to_image_path=None):
    '''
    将paddleocr识别出来的矩形框画到图像上。
    '''
    try:
        result = json_data[0]
        image = Image.open(image_path).convert('RGB')
        boxes = [line[0] for line in result]
        txts = [line[1][0] for line in result]
        scores = [line[1][1] for line in result]
        im_show = draw_ocr(image, boxes, font_path='./fonts/simfang.ttf')
        im_show = Image.fromarray(im_show)
        if save_to_image_path is not None:
            im_show.save(save_to_image_path)
    except Exception as e:
        print(f"An error occurred: {e}")

def check_ocr_file_existence(filenames):
    '''
    检查文件的 ocr 识别处理文件是否存在。
    filenames 是 ocr_build_other_filepath 的返回值。
    '''
    # 初始化标志变量为默认值
    json_log_flag = 0
    table_log_flag = 0
    cell_log_flag = 0
    content_log_flag = 0
    table_content_log_flag = 0
    json_size_flag = 0

    # 检查文件名列表是否为空
    if not filenames:
        return filenames[0], filenames[1], json_log_flag, table_log_flag, cell_log_flag, content_log_flag, table_content_log_flag, json_size_flag

    # 检查每个文件名对应的文件是否存在
    try:
        check_json = check_file_existence(filenames[1])
        check_table = check_file_existence(filenames[2])
        check_cell = check_file_existence(filenames[3])
        check_content = check_file_existence(filenames[4])
        check_table_content = check_file_existence(filenames[6])
    except IndexError:
        # 如果索引超出范围，说明文件名列表不完整，直接返回默认值
        return filenames[0], filenames[1], json_log_flag, table_log_flag, cell_log_flag, content_log_flag, table_content_log_flag, json_size_flag

    # 如果文件存在，进行后续处理
    if check_json:
        json_log_flag = 1
        # 尝试获取文件大小（以字节为单位）
        try:
            file_size = os.path.getsize(filenames[1])
            # 10KB 的字节数
            paddlejson_file_limit_kb = 10 * 1024
            if file_size <= paddlejson_file_limit_kb:
                json_size_flag = 0
            else:
                json_size_flag = 1
        except OSError:
            # 如果获取文件大小出错，设置标志为 0
            json_size_flag = 0

    table_log_flag = 1 if check_table else 0
    cell_log_flag = 1 if check_cell else 0
    content_log_flag = 1 if check_content else 0
    table_content_log_flag = 1 if check_table_content else 0

    return filenames[0], filenames[1], json_log_flag, table_log_flag, cell_log_flag, content_log_flag, table_content_log_flag, json_size_flag

def check_ocr_result_save_to_excel(filepaths, ocr_model='ali'):
    '''
    filenames的格式如下：
    ori_filename, json_filename, table_filename, cell_filename, content_filename, log_filename, table_content_filename
    '''
    column_names = ['fileName', 'jsonName','json_flag', 'table_flag', 'cell_flag', 'content_flag', 'table_content_flag','json_size_flag']
    df = pd.DataFrame()
    for f in filepaths:
        filenames = ocr_build_other_filepath(f, ocr_model)
        result = check_ocr_file_existence(filenames)
        temp_df = pd.DataFrame([result], columns=column_names)
        df = pd.concat([df, temp_df], ignore_index=True)
    return df



