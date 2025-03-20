import argparse
import datetime
import os
import openpyxl

def get_all_possible_paths(dir_path, exclude):
    try:
        all_paths = []
        for dirpath, dirnames, _ in os.walk(dir_path):
            for dirname in dirnames:
                all_paths.append(os.path.join(dirpath, dirname))
        
        filtered_paths = [item for item in all_paths if not any(excl in item for excl in exclude)]
        filtered_paths = sorted(filtered_paths)

        result_folder = []
        for item in filtered_paths:
            subdirs = [d for d in os.listdir(item) if os.path.isdir(os.path.join(item, d))]
            if not subdirs:
                result_folder.append(item)
        return result_folder
    except Exception as e:
        print(f"An error occurred: {e}")

def write_to_txt(file_name, all_path):
    with open(f"{file_name}.txt", "w", encoding="utf-8") as f:
        for path in all_path:
            f.write(f"{path}\n")
def write_to_excel(directory, file_name, all_path):    
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="ลำดับ")
        ws.cell(row=1, column=2, value="โฟลเดอร์")
        for row_idx, path in enumerate(all_path, start=2):
            path = path.replace(directory, "")
            ws.cell(row=row_idx, column=1, value=row_idx-1)
            ws.cell(row=row_idx, column=2, value=path)
        wb.save(f'{file_name}.xlsx')
    except Exception as e:
        print(f"ERROR: {e}")

def tree(directory, exclude, file_name):               
    all_path = get_all_possible_paths(directory, exclude)
    write_to_txt(file_name, all_path)
    write_to_excel(directory, file_name, all_path)

def parse_args():
    parser = argparse.ArgumentParser(description="Generate a directory tree listing.")
    parser.add_argument("-i", "--input_folder", type=str, default=".", help="Input folder to start the tree generation.")
    return parser.parse_args()

def read_exclude_file():
    exclude_file = ".exclude"
    if os.path.exists(exclude_file):
        with open(exclude_file, "r", encoding="utf-8") as f:
            return [line.strip() for line in f.readlines() if line.strip()]
    return []

def main():
    args = parse_args()
    file_name = datetime.datetime.now().strftime('%Y-%m-%d')
    exclude_list = read_exclude_file() + [".exclude", f"{file_name}.txt", f"{file_name}.xlsx"]
    tree(args.input_folder, exclude=exclude_list, file_name=file_name)
    print(f"Tree structure generated and saved to {file_name}.txt and {file_name}.xlsx")

if __name__ == "__main__":
    main()
