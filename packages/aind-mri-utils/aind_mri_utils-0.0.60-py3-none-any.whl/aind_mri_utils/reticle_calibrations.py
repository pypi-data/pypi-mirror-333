"""
Functions to read reticle calibration data, find a transformation between
coordinate frames, and apply the transformation.
"""

import io

import numpy as np
from openpyxl import load_workbook
from scipy import optimize as opt
from scipy.spatial.transform import Rotation

from . import rotations as rot


def extract_calibration_metadata(ws):
    """
    Extract calibration metadata from an Excel worksheet.

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        The worksheet object from which to extract the calibration metadata.

    Returns
    -------
    tuple
        A tuple containing:
        - global_factor (float): The global scale value.
        - global_rotation_degrees (float): The global rotation in degrees.
        - manipulator_factor (float): The manipulator scale value.
        - global_offset (numpy.ndarray): The global offset as a 3-element
          array.
        - reticle_name (str): The name of the reticle.
    """
    row_iter = ws.iter_rows(min_row=1, max_row=2, values_only=True)
    col_name_lookup = {k: i for i, k in enumerate(next(row_iter))}
    metadata_values = next(row_iter)
    global_factor = metadata_values[col_name_lookup["GlobalFactor"]]
    global_rotation_degrees = metadata_values[
        col_name_lookup["GlobalRotationDegrees"]
    ]
    manipulator_factor = metadata_values[col_name_lookup["ManipulatorFactor"]]
    reticle_name = metadata_values[col_name_lookup["Reticule"]]
    offset_x_pos = col_name_lookup["GlobalOffsetX"]
    global_offset = np.array(
        metadata_values[offset_x_pos : offset_x_pos + 3],  # noqa: E203
        dtype=float,
    )
    return (
        global_factor,
        global_rotation_degrees,
        manipulator_factor,
        global_offset,
        reticle_name,
    )


def _contains_none(arr):
    """Checks if all arguments are not None."""
    return any(x is None for x in arr)


def extract_calibration_pairs(ws):
    """
    Extract calibration pairs from an Excel worksheet.

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        The worksheet object from which to extract the calibration pairs.

    Returns
    -------
    dict
        A dictionary where keys are probe names and values are lists of tuples,
        each containing a reticle point and a probe point as numpy arrays.
    """
    pairs_by_probe = dict()
    for row in ws.iter_rows(min_row=2, max_col=7, values_only=True):
        probe_name = row[0]
        if probe_name is None:
            continue
        reticle_pt = np.array(row[1:4])
        probe_pt = np.array(row[4:7])
        if _contains_none(reticle_pt) or _contains_none(probe_pt):
            continue
        if probe_name not in pairs_by_probe:
            pairs_by_probe[probe_name] = []
        pairs_by_probe[probe_name].append((reticle_pt, probe_pt))
    return pairs_by_probe


def _combine_pairs(list_of_pairs):
    """
    Combine lists of pairs into separate global and manipulator points
    matrices.

    Parameters
    ----------
    list_of_pairs : list of tuple
        A list of tuples, each containing a reticle point and a probe point as
        numpy arrays.

    Returns
    -------
    tuple
        Two numpy arrays, one for global points and one for manipulator points.
    """
    global_pts, manipulator_pts = [np.vstack(x) for x in zip(*list_of_pairs)]
    return global_pts, manipulator_pts


def _apply_metadata_to_pair_mats(
    global_pts,
    manipulator_pts,
    global_factor,
    global_rotation_degrees,
    global_offset,
    manipulator_factor,
):
    """
    Apply calibration metadata to global and manipulator points matrices.

    Parameters
    ----------
    global_pts : numpy.ndarray
        The global points matrix.
    manipulator_pts : numpy.ndarray
        The manipulator points matrix.
    global_factor : float
        The global factor value.
    global_rotation_degrees : float
        The global rotation in degrees.
    global_offset : numpy.ndarray
        The global offset as a 3-element array.
    manipulator_factor : float
        The manipulator factor value.

    Returns
    -------
    tuple
        The adjusted global points and manipulator points matrices.
    """
    if global_rotation_degrees != 0:
        rot_mat = (
            Rotation.from_euler("z", global_rotation_degrees, degrees=True)
            .as_matrix()
            .squeeze()
        )
        # Transposed because points are row vectors
        global_pts = global_pts @ rot_mat.T
    global_pts = global_pts * global_factor + global_offset
    manipulator_pts = manipulator_pts * manipulator_factor
    return global_pts, manipulator_pts


def _apply_metadata_to_pair_lists(
    list_of_pairs,
    global_factor,
    global_rotation_degrees,
    global_offset,
    manipulator_factor,
):
    """
    Apply calibration metadata to lists of pairs.

    Parameters
    ----------
    list_of_pairs : list of tuple
        A list of tuples, each containing a reticle point and a probe point as
        numpy arrays.
    global_factor : float
        The global factor value.
    global_rotation_degrees : float
        The global rotation in degrees.
    global_offset : numpy.ndarray
        The global offset as a 3-element array.
    manipulator_factor : float
        The manipulator factor value.

    Returns
    -------
    tuple
        The adjusted global points and manipulator points matrices.
    """
    global_pts, manipulator_pts = _combine_pairs(list_of_pairs)
    return _apply_metadata_to_pair_mats(
        global_pts,
        manipulator_pts,
        global_factor,
        global_rotation_degrees,
        global_offset,
        manipulator_factor,
    )


def read_reticle_calibration(
    filename, points_sheet_name="points", metadata_sheet_name="metadata"
):
    """
    Read reticle calibration data from an Excel file.

    Parameters
    ----------
    filename : str
        The path to the Excel file containing the calibration data.
    points_sheet_name : str, optional
        The name of the sheet containing the calibration points.
        The default is "points".
    metadata_sheet_name : str, optional
        The name of the sheet containing the calibration metadata.
        The default is "metadata".

    Returns
    -------
    tuple
        A tuple containing:
        - adjusted_pairs_by_probe (dict): Adjusted calibration pairs by probe
          name.
        - global_offset (numpy.ndarray): The global offset as a 3-element
          array.
        - global_rotation_degrees (float): The global rotation in degrees.
        - reticle_name (str): The name of the reticle.

    Raises
    ------
    ValueError
        If the specified sheets are not found in the Excel file.
    """
    in_mem_file = None
    with open(filename, "rb") as f:
        in_mem_file = io.BytesIO(f.read())
    wb = load_workbook(in_mem_file, read_only=True, data_only=True)
    if points_sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet {points_sheet_name} not found in {filename}")
    if metadata_sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet {metadata_sheet_name} not found in {filename}"
        )
    (
        global_factor,
        global_rotation_degrees,
        manipulator_factor,
        global_offset,
        reticle_name,
    ) = extract_calibration_metadata(wb[metadata_sheet_name])
    pairs_by_probe = extract_calibration_pairs(wb["points"])
    adjusted_pairs_by_probe = {
        k: _apply_metadata_to_pair_lists(
            v,
            global_factor,
            global_rotation_degrees,
            global_offset,
            manipulator_factor,
        )
        for k, v in pairs_by_probe.items()
    }
    return (
        adjusted_pairs_by_probe,
        global_offset,
        global_rotation_degrees,
        reticle_name,
    )


def _unpack_theta(theta):
    """Helper function to unpack theta into rotation matrix and translation."""
    R = rot.combine_angles(*theta[0:3])
    offset = theta[3:6]
    return R, offset


def fit_rotation_params(
    reticle_pts, probe_pts, legacy_outputs=False, find_scaling=True, **kwargs
):
    """
    Fit rotation parameters to align reticle points with probe points using
    least squares optimization. The rotation matrix and translation vector
    are the solution for the equation

    probe_pts = R @ reticle_pts + translation

    where each point is a column vector.

    Because numpy is row-major, points are often stored as row vectors. In this
    case, you should use the transpose of this equation:

    probe_pts = reticle_pts @ R.T + translation

    Parameters
    ----------
    reticle_pts : numpy.ndarray
        The reticle points to be transformed.
    probe_pts : numpy.ndarray
        The probe points to align with.
    legacy_outputs : bool, optional
        If True, return the translation in the global frame and the transpose
        of the rotation matrix.  The default is False.
    **kwargs : dict
        Additional keyword arguments to pass to the least squares optimization
        function.

    Returns
    -------
    tuple
        A tuple containing:
        - R (numpy.ndarray): The 3x3 rotation matrix.
        - translation (numpy.ndarray): The 3-element translation vector.
    """
    if reticle_pts.shape != probe_pts.shape:
        raise ValueError("reticle_pts and probe_pts must have the same shape")
    if reticle_pts.shape[1] != 3:
        raise ValueError("reticle_pts and probe_pts must have 3 columns")

    if find_scaling:
        if legacy_outputs:
            raise NotImplementedError(
                "find_scaling=True not valid when legacy_outputs=True"
            )
        return _fit_params_with_scaling(reticle_pts, probe_pts, **kwargs)

    R_homog = np.eye(4)
    reticle_pts_homog = rot.prepare_data_for_homogeneous_transform(reticle_pts)
    transformed_pts_homog = np.empty_like(reticle_pts_homog)

    def fun(theta):
        """cost function for least squares optimization"""
        R_homog[0:3, 0:3] = rot.combine_angles(*theta[0:3])
        R_homog[0:3, 3] = theta[3:6]  # translation
        np.matmul(reticle_pts_homog, R_homog.T, out=transformed_pts_homog)
        residuals = (transformed_pts_homog[:, 0:3] - probe_pts).flatten()
        return residuals

    # Initial guess of parameters
    theta0 = np.zeros(6)

    if probe_pts.shape[0] > 1:
        # Initial guess of rotation: align the vectors between the first
        # two points
        probe_diff = np.diff(probe_pts[:2, :], axis=0)
        reticle_diff = np.diff(reticle_pts[:2, :], axis=0)
        R_init = rot.rotation_matrix_from_vectors(
            reticle_diff.squeeze(), probe_diff.squeeze()
        )
        theta0[0:3] = Rotation.from_matrix(R_init).as_euler("xyz")

    # Initial guess of translation: find the point on the reticle closest to
    # zero
    smallest_pt = np.argmin(np.linalg.norm(reticle_pts, axis=1))
    theta0[3:6] = probe_pts[smallest_pt, :]

    res = opt.least_squares(fun, theta0, **kwargs)
    R, translation = _unpack_theta(res.x)
    if legacy_outputs:
        # last version had translation in global frame
        #
        # Also the last version for some reason calculated the transpose of the
        # rotation matrix. Application of the rotation matrix was consistently
        # wrong in a way that accounted for this transpose, so the results were
        # correct.
        #
        # All of the transposes are confusing here: this is the inverse of the
        # rotation matrix, accounting for numpy being row-major, and
        # data points being row vectors

        translation = translation @ R  # Not R.T!
        return translation, R.T  # Not R!
    scaling = None
    return R, translation, scaling


def fit_rotation_params_from_excel(filename, *args, **kwargs):
    adjusted_pairs_by_probe = read_reticle_calibration(filename)[0]
    cal_by_probe = {
        k: fit_rotation_params(*v, *args, **kwargs)
        for k, v in adjusted_pairs_by_probe.items()
    }
    return cal_by_probe


def _unpack_theta_scale(theta):
    """Helper function to unpack theta into rotation matrix and translation."""
    R = rot.combine_angles(*theta[0:3])
    scale = theta[3:6]
    translation = theta[6:]
    return R, scale, translation


def _fit_params_with_scaling(reticle_pts, probe_pts, **kwargs):
    """
    Fit rotation parameters to align reticle points with probe points using
    least squares optimization. The rotation matrix and translation vector
    are the solution for the equation

    probe_pts = R @ reticle_pts + translation

    where each point is a column vector.

    Because numpy is row-major, points are often stored as row vectors. In this
    case, you should use the transpose of this equation:

    probe_pts = reticle_pts @ R.T + translation

    Parameters
    ----------
    reticle_pts : numpy.ndarray
        The reticle points to be transformed.
    probe_pts : numpy.ndarray
        The probe points to align with.
    find_scaling : bool, optional
        If True, find a scaling factor to apply to the reticle points.
        The default is True.
    **kwargs : dict
        Additional keyword arguments to pass to the least squares optimization
        function.

    Returns
    -------
    tuple
        A tuple containing:
        - R (numpy.ndarray): The 3x3 rotation matrix.
        - translation (numpy.ndarray): The 3-element translation vector.
        - scaling (float): The scaling factor.
    """

    def fun(theta):
        """cost function for least squares optimization"""
        R, scale, translation = _unpack_theta_scale(theta)
        transformed_reticle = reticle_pts @ R.T * scale + translation
        residuals = (transformed_reticle - probe_pts).flatten()
        return residuals

    # Initial guess of parameters
    theta0 = np.zeros(9)
    theta0[3:6] = 1.0
    npt = probe_pts.shape[0]
    if npt > 1:
        # Initial guess of rotation: align the vectors between the first
        # two points
        for other_pt in range(1, npt):
            probe_diff = probe_pts[other_pt, :] - probe_pts[0, :]
            reticle_diff = reticle_pts[other_pt, :] - reticle_pts[0, :]
            reticle_norm = np.linalg.norm(reticle_diff.squeeze())
            if reticle_norm > 0:
                break
        if reticle_norm > 0:
            R_init = rot.rotation_matrix_from_vectors(
                reticle_diff.squeeze(), probe_diff.squeeze()
            )
            theta0[0:3] = Rotation.from_matrix(R_init).as_euler("xyz")

    # Initial guess of translation: find the point on the reticle closest to
    # zero
    smallest_pt = np.argmin(np.linalg.norm(reticle_pts, axis=1))
    theta0[6:] = probe_pts[smallest_pt, :]

    res = opt.least_squares(fun, theta0, **kwargs)
    R, scale, translation = _unpack_theta_scale(res.x)
    return R, translation, scale


def _apply_scale_to_rotation(R, scale):
    """
    Apply a scaling factor to a rotation matrix.

    Parameters
    ----------
    R : numpy.ndarray
        The 3x3 rotation matrix.
    scale : float
        The scaling factor.

    Returns
    -------
    numpy.ndarray
        The scaled rotation matrix.
    """
    scale_mat = np.zeros((3, 3))
    np.fill_diagonal(scale_mat, scale)
    return scale_mat @ R


def transform_reticle_to_probe(reticle_pts, R, translation, scale=None):
    """
    Transform reticle points to probe points using rotation and translation.

    Parameters
    ----------
    probe_pts : np.array(N,3)
        Probe points to transform.
    R : np.array(3,3)
        Rotation matrix.
    translation : np.array(3,)
        Translation vector.

    Returns
    -------
    np.array(N,3)
        Transformed points.
    """
    if scale is None:
        transformed = rot.apply_rotate_translate(reticle_pts, R, translation)
    else:
        transformed = rot.apply_rotate_translate_scale(
            reticle_pts, R, translation, scale
        )
    return transformed


def transform_probe_to_reticle(probe_pts, R, translation, scale=None):
    """
    Transform probe points to reticle points using rotation and translation.

    Parameters
    ----------
    probe_pts : np.array(N,3)
        Probe points to transform.
    R : np.array(3,3)
        Rotation matrix.
    translation : np.array(3,)
        Translation vector.

    Returns
    -------
    np.array(N,3)
        Transformed points.
    """
    if scale is None:
        R_inv, t_inv = rot.inverse_rotate_translate(R, translation)
        transformed = rot.apply_rotate_translate(probe_pts, R_inv, t_inv)
    if scale is not None:
        transformed = rot.apply_inverse_rotate_translate_scale(
            probe_pts, R, translation, scale
        )
    return transformed
